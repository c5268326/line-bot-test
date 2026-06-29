"""
用 IMAP 抓取最新業績 Excel，解析後更新 performance.json，並廣播推播最新業績
"""
import imaplib
import email
from email.header import decode_header
import os
import json
import io
import re
import requests
from datetime import datetime, timezone, timedelta
import openpyxl
import xlrd

GMAIL_USER = os.environ.get("GMAIL_USER", "wangnanshan33@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get(
    "LINE_CHANNEL_ACCESS_TOKEN",
    "SFqZWlkhtiRJXJNHgjO6PEqeHbmgIq3Ww2fO5kiq/25W+8CjFF9bApG9e9/VzuAJSZlPsSs/VUEFWAos4nyKOzAihgrzfkjCz8kxcb7w7ogiw01htnA65RIziuKn/hlaVCjwZCu8orjs0IH0hxY1ZQdB04t89/1O/w1cDnyilFU="
)

DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "performance.json")

REGIONS = ["全國", "台北一區", "台北二區", "桃竹苗區", "中部地區", "南部地區"]

DEPARTMENTS = {
    "台北一區": ["北一業展一處", "北一業展二處", "北一業展三處", "北一業展四處", "北一業展五處", "北一業展六處", "北一業展七處"],
    "台北二區": ["北二業展一處", "北二業展二處", "北二業展三處", "北二業展四處", "北二業展五處", "北二業展六處", "北二業展七處"],
    "桃竹苗區": ["桃竹苗業展一處", "桃竹苗業展二處", "桃竹苗業展三處", "桃竹苗業展四處"],
    "中部地區": ["中區業展一處", "中區業展二處", "中區業展三處", "中區業展四處", "中區業展五處", "中區業展六處", "中區業展七處"],
    "南部地區": ["南區業展一處", "南區業展二處", "南區業展三處", "南區業展四處"],
}
ALL_DEPTS = [d for depts in DEPARTMENTS.values() for d in depts]
ALL_NAMES = set(REGIONS) | set(ALL_DEPTS)


def _get_mail_ids(mail):
    """搜尋指定寄件者信件，找不到則搜全部"""
    SENDERS = [
        "Jessica-YC.Lu@nanshan.com.tw",
        "c5268326@gmail.com",
    ]
    _, data = mail.search(None, f'(OR FROM "{SENDERS[0]}" FROM "{SENDERS[1]}")')
    mail_ids = data[0].split()
    if not mail_ids:
        _, data = mail.search(None, "ALL")
        mail_ids = data[0].split()
    return mail_ids


def _decode_filename(raw_filename):
    decoded_parts = decode_header(raw_filename)
    filename = ""
    for part_bytes, charset in decoded_parts:
        if isinstance(part_bytes, bytes):
            filename += part_bytes.decode(charset or "utf-8", errors="replace")
        else:
            filename += part_bytes
    return filename


def _decode_subject(msg):
    """解碼信件主旨"""
    raw = msg.get("Subject", "")
    parts = decode_header(raw)
    subject = ""
    for part_bytes, charset in parts:
        if isinstance(part_bytes, bytes):
            subject += part_bytes.decode(charset or "utf-8", errors="replace")
        else:
            subject += part_bytes
    return subject


def _get_excel_attachment(msg):
    """取得信件中第一個 Excel 附件的 (BytesIO, filename)，無則回傳 (None, None)"""
    for part in msg.walk():
        raw_filename = part.get_filename()
        if not raw_filename:
            continue
        filename = _decode_filename(raw_filename)
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            return io.BytesIO(part.get_payload(decode=True)), filename
    return None, None


def get_latest_excels():
    """連線 Gmail IMAP，依信件主旨分別找最新的月報表和日報表
    - 日報表：主旨含「速報」
    - 月報表：主旨含「每日業績追蹤報表」
    """
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")

    mail_ids = _get_mail_ids(mail)
    print(f"搜尋到 {len(mail_ids)} 封信")

    monthly_file = None
    monthly_name = None
    today_file = None
    today_name = None

    for mail_id in reversed(mail_ids):
        if monthly_file and today_file:
            break
        _, msg_data = mail.fetch(mail_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        subject = _decode_subject(msg)
        sender = msg.get("From", "")
        print(f"信件：{subject}（{sender}）")

        if "速報" in subject and today_file is None:
            file_bytes, filename = _get_excel_attachment(msg)
            if file_bytes:
                today_file = file_bytes
                today_name = filename
                print(f"✅ 日報表：{filename}（主旨：{subject}）")

        elif "每日業績追蹤報表" in subject and monthly_file is None:
            file_bytes, filename = _get_excel_attachment(msg)
            if file_bytes:
                monthly_file = file_bytes
                monthly_name = filename
                print(f"✅ 月報表：{filename}（主旨：{subject}）")

    mail.logout()
    return (monthly_file, monthly_name), (today_file, today_name)


def _parse_time_from_filename(filename):
    """從檔名解析時間（取整點），如 _20260629_2127.xlsx → 2026/06/29 21:00"""
    m = re.search(r'(\d{8})_(\d{4})', filename)
    if m:
        d, t = m.group(1), m.group(2)
        return f"{d[:4]}/{d[4:6]}/{d[6:8]} {t[:2]}:00"
    return None


def parse_today_excel(file_bytes, filename=""):
    """
    解析日報表（7欄格式）：
      A欄 = 地區/業展處, B=實收保費, C=實收達成率, D=A&H保費, E=A&H達成率, F=RP保費, G=RP達成率
      row 0 = 標題, row 1 起為資料
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))

    TW = timezone(timedelta(hours=8))
    report_time = _parse_time_from_filename(filename) or datetime.now(TW).strftime("%Y/%m/%d %H:%M")
    print(f"📅 日報表時間：{report_time}")

    today_regions = {}
    today_depts = {}

    for row in all_rows:
        if not row[0]:
            continue
        raw = str(row[0]).strip()
        sys_name = REGION_NAME_MAP.get(raw)

        def v(i):
            val = row[i] if len(row) > i else None
            return str(val) if val not in (None, "") else "－"

        vals = {"實收保費": v(1), "實收達成率": v(2), "A&H保費": v(3), "A&H達成率": v(4), "RP保費": v(5), "RP達成率": v(6)}

        if sys_name:
            today_regions[sys_name] = vals
        elif raw in REGIONS:
            today_regions[raw] = vals
        else:
            for region, depts in DEPARTMENTS.items():
                if raw in depts:
                    today_depts.setdefault(region, {})[raw] = vals
                    break

    print(f"✅ 日報表：地區 {len(today_regions)} 筆，業展處 {sum(len(v) for v in today_depts.values())} 筆")
    return today_regions, today_depts, report_time


# 新格式：Excel 地區名稱 → 系統地區名稱
REGION_NAME_MAP = {
    "北一業發部": "台北一區",
    "北二業發部": "台北二區",
    "桃竹苗業發部": "桃竹苗區",
    "中區業發部": "中部地區",
    "南區業發部": "南部地區",
    "合計": "全國",
}

# 新格式：各地區業展處在 all_rows 中的起始列索引與數量（0-based）
# 地區列：rows 2-7（0-based），業展處：rows 11-39（0-based）
DEPT_ROW_MAP = {
    "台北一區":  (11, 7),
    "台北二區":  (18, 7),
    "桃竹苗區":  (25, 4),
    "中部地區":  (29, 7),
    "南部地區":  (36, 4),
}


def _extract_new_format(row, today=False):
    """從新格式列擷取業績欄位"""
    def v(i):
        val = row[i] if len(row) > i else None
        return str(val) if val not in (None, "") else "－"

    if today:
        # col 8=當日實收, 9=當日實收率, 10=當日A&H, 11=當日A&H率, 12=當日RP, 13=當日RP率
        return {"實收保費": v(8), "實收達成率": v(9), "A&H保費": v(10), "A&H達成率": v(11), "RP保費": v(12), "RP達成率": v(13)}
    else:
        # col 17=本月實收, 18=本月實收率, 23=A&H, 24=A&H率, 27=期繳, 28=期繳率
        return {"實收保費": v(17), "實收達成率": v(18), "A&H保費": v(23), "A&H達成率": v(24), "RP保費": v(27), "RP達成率": v(28)}


def _parse_report_time(all_rows):
    """從 row0 col14 的標題文字解析報表時間，如：6/26 11點 → 2026/06/26 11:00"""
    try:
        header = str(all_rows[0][14]) if len(all_rows[0]) > 14 else ""
        m = re.search(r'(\d+)/(\d+)\s+(\d+)點', header)
        if m:
            TW = timezone(timedelta(hours=8))
            year = datetime.now(TW).year
            month, day, hour = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return f"{year}/{month:02d}/{day:02d} {hour:02d}:00"
    except Exception:
        pass
    return None


def _parse_new_format(all_rows):
    """解析新版單工作表格式（業發部_業展處_三時段整點）"""
    monthly = {}
    today = {}

    # 地區列 rows 2-7（含合計）
    for row in all_rows[2:8]:
        raw = str(row[0]).strip() if row[0] else ""
        sys_name = REGION_NAME_MAP.get(raw)
        if sys_name:
            monthly[sys_name] = _extract_new_format(row, today=False)
            today[sys_name] = _extract_new_format(row, today=True)
            print(f"✅ 地區：{raw} → {sys_name}")

    # 業展處列（依位置對應部門名稱）
    for region, (start_idx, count) in DEPT_ROW_MAP.items():
        dept_names = DEPARTMENTS[region]
        for i, row in enumerate(all_rows[start_idx:start_idx + count]):
            if i >= len(dept_names):
                break
            dept_name = dept_names[i]
            monthly[dept_name] = _extract_new_format(row, today=False)
            today[dept_name] = _extract_new_format(row, today=True)
            print(f"✅ 業展處：{region}[{i+1}] → {dept_name}")

    report_time = _parse_report_time(all_rows)
    print(f"📅 報表時間：{report_time}")
    return monthly, today, report_time


def _split_monthly(flat_monthly):
    """將 _parse_new_format 回傳的 flat dict 拆分為 regions 和 depts"""
    regions = {}
    depts = {}
    for name, vals in flat_monthly.items():
        if name in REGIONS:
            regions[name] = vals
        else:
            for region, dept_list in DEPARTMENTS.items():
                if name in dept_list:
                    depts.setdefault(region, {})[name] = vals
                    break
    return regions, depts


def parse_xls(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes.read())
    all_rows = [wb.sheet_by_index(0).row_values(i) for i in range(wb.sheet_by_index(0).nrows)]
    flat, _, report_time = _parse_new_format(all_rows)
    regions, depts = _split_monthly(flat)
    return regions, depts, report_time


def parse_excel(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
    print(f"工作表：{wb.worksheets[0].title}，共 {len(all_rows)} 列")
    flat, _, report_time = _parse_new_format(all_rows)
    regions, depts = _split_monthly(flat)
    return regions, depts, report_time


def _combine_vals(monthly_vals, today_vals):
    """合併月報（月初到昨日17:00）+ 日報（昨日17:00到現在）的金額，重新計算達成率"""
    def to_float(v):
        try:
            return float(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return None

    combined = {}
    for amt_key, rate_key in [("實收保費", "實收達成率"), ("A&H保費", "A&H達成率"), ("RP保費", "RP達成率")]:
        m_amt = to_float(monthly_vals.get(amt_key))
        d_amt = to_float(today_vals.get(amt_key))
        m_rate = parse_rate_float(monthly_vals.get(rate_key))

        if m_amt is not None and d_amt is not None:
            combined_amt = m_amt + d_amt
            combined[amt_key] = str(combined_amt)
            # 用月報達成率反推目標，計算合併達成率
            if m_rate and m_rate > 0 and m_amt > 0:
                target = m_amt / m_rate
                combined[rate_key] = f"{combined_amt / target * 100:.1f}%"
            else:
                combined[rate_key] = monthly_vals.get(rate_key, "－")
        else:
            combined[amt_key] = monthly_vals.get(amt_key, "－")
            combined[rate_key] = monthly_vals.get(rate_key, "－")

    return combined


def update_performance(monthly=None, monthly_depts=None, today_regions=None, today_depts=None, report_time=None):
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "departments" not in data:
        data["departments"] = {r: {} for r in DEPARTMENTS}

    # 更新本月地區（月報 + 日報合併）
    if monthly:
        for name, values in monthly.items():
            if today_regions and name in today_regions:
                data["regions"][name] = _combine_vals(values, today_regions[name])
                print(f"✅ 本月地區合併 {name}")
            else:
                data["regions"][name] = values
                print(f"✅ 本月地區更新 {name}")

    # 更新本月業展處（月報 + 日報合併）
    if monthly_depts:
        for region, depts in monthly_depts.items():
            for dept, values in depts.items():
                today_dept_vals = (today_depts or {}).get(region, {}).get(dept)
                if today_dept_vals:
                    data["departments"].setdefault(region, {})[dept] = _combine_vals(values, today_dept_vals)
                    print(f"✅ 本月業展處合併 {dept}")
                else:
                    data["departments"].setdefault(region, {})[dept] = values
                    print(f"✅ 本月業展處更新 {dept}")

    # 更新今日（保留昨日）
    if today_regions or today_depts:
        if data.get("today"):
            data["yesterday"] = data["today"]
        if data.get("today_departments"):
            data["yesterday_departments"] = data["today_departments"]
        data["today"] = today_regions or {}
        data["today_departments"] = today_depts or {r: {} for r in DEPARTMENTS}

    if report_time:
        data["updated_at"] = report_time
    else:
        TW = timezone(timedelta(hours=8))
        data["updated_at"] = datetime.now(TW).strftime("%Y/%m/%d %H:%M")

    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("performance.json 已更新")


def fmt_amount(val):
    try:
        n = float(str(val).replace(",", ""))
        return f"{n:,.0f}"
    except (ValueError, TypeError):
        return val


def fmt_rate(val):
    s = str(val).strip()
    if s in ("－", "", "None"):
        return "－"
    if "%" in s:
        try:
            return f"{float(s.replace('%', '')):.1f}%"
        except ValueError:
            return s
    try:
        n = float(s)
        if n < 20:
            return f"{n * 100:.1f}%"
        return f"{n:.1f}%"
    except ValueError:
        return s


def parse_rate_float(val):
    s = str(val).strip().replace("%", "")
    try:
        n = float(s)
        return n / 100 if n >= 20 else n
    except ValueError:
        return None


def broadcast_performance():
    """廣播最新業績 Flex Message 給所有使用者"""
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 本日廣播優先用日報表資料（today），無則退回月報表
    regions = data.get("today") or data.get("regions", {})
    updated = data.get("updated_at", "－")
    national = regions.get("全國", {})

    def rate_color(val, nat_key):
        f = parse_rate_float(val)
        nf = parse_rate_float(national.get(nat_key, "0"))
        if f is not None and nf is not None and f < nf:
            return "#e74c3c"
        return "#111111"

    def make_row(label, amount, rate_val, nat_key, is_national=False):
        color = "#111111" if is_national else rate_color(rate_val, nat_key)
        return {
            "type": "box", "layout": "horizontal",
            "contents": [
                {"type": "text", "text": label, "size": "sm", "color": "#666666", "flex": 2},
                {"type": "text", "text": fmt_amount(amount), "size": "sm", "color": "#222222", "flex": 4, "align": "end", "weight": "bold"},
                {"type": "text", "text": fmt_rate(rate_val), "size": "sm", "color": color, "flex": 3, "align": "end", "weight": "bold"},
            ],
            "margin": "xs",
        }

    blocks = []
    for region, vals in regions.items():
        is_nat = (region == "全國")
        blocks.append({"type": "text", "text": f"【{region}】", "weight": "bold", "size": "md", "color": "#1a5276", "margin": "md"})
        blocks.append(make_row("實收", vals.get("實收保費", "－"), vals.get("實收達成率", "－"), "實收達成率", is_nat))
        blocks.append(make_row("A&H", vals.get("A&H保費", "－"), vals.get("A&H達成率", "－"), "A&H達成率", is_nat))
        blocks.append(make_row("RP", vals.get("RP保費", "－"), vals.get("RP達成率", "－"), "RP達成率", is_nat))
        blocks.append({"type": "separator", "margin": "md"})

    flex_message = {
        "type": "flex",
        "altText": "📊 最新業績更新通知",
        "contents": {
            "type": "bubble",
            "size": "mega",
            "header": {
                "type": "box", "layout": "vertical",
                "contents": [
                    {"type": "text", "text": "📊 最新業績更新通知", "weight": "bold", "size": "lg", "color": "#1a5276"},
                    {"type": "text", "text": f"截至 {updated}", "size": "xs", "color": "#888888", "margin": "xs"},
                ],
                "backgroundColor": "#EBF5FB",
                "paddingAll": "16px",
            },
            "body": {
                "type": "box", "layout": "vertical",
                "contents": blocks,
                "paddingAll": "12px",
                "spacing": "none",
            },
        },
    }

    resp = requests.post(
        "https://api.line.me/v2/bot/message/broadcast",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}",
        },
        json={"messages": [flex_message]},
    )
    if resp.status_code == 200:
        print("✅ 廣播推播成功")
    else:
        print(f"⚠️ 廣播推播失敗：{resp.status_code} {resp.text}")


def main():
    if not GMAIL_APP_PASSWORD:
        print("❌ 未設定 GMAIL_APP_PASSWORD 環境變數")
        return

    (monthly_file, monthly_name), (today_file, today_name) = get_latest_excels()

    monthly_regions = {}
    monthly_depts = {}
    report_time = None

    if monthly_file:
        monthly_regions, monthly_depts, report_time = parse_excel(monthly_file)
    else:
        print("⚠️ 找不到月報表（49欄格式）")

    today_regions = {}
    today_depts = {}
    today_time = None

    if today_file:
        today_regions, today_depts, today_time = parse_today_excel(today_file, today_name or "")
    else:
        print("⚠️ 找不到日報表（7欄格式）")

    if not monthly_regions and not today_regions:
        print("⚠️ 兩份報表皆無資料")
        return

    # 以日報表時間為主，月報表時間為輔
    final_time = today_time or report_time

    update_performance(
        monthly=monthly_regions,
        monthly_depts=monthly_depts,
        today_regions=today_regions,
        today_depts=today_depts,
        report_time=final_time,
    )
    broadcast_performance()


if __name__ == "__main__":
    main()
